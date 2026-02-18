#!/usr/bin/env python3
"""
Create Excel spreadsheet template for opponent difficulty classification
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_opponent_difficulty_spreadsheet():
    print("📊 Creating Opponent Difficulty Classification Spreadsheet...")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet and create our sheets
    wb.remove(wb.active)
    
    # Create main data entry sheet
    ws_main = wb.create_sheet("Opponent Analysis")
    
    # Create instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    
    # Create examples sheet
    ws_examples = wb.create_sheet("Examples")
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, color="000000")
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # === MAIN DATA ENTRY SHEET ===
    print("Creating main data entry sheet...")
    
    # Headers
    ws_main['A1'] = "OPPONENT DIFFICULTY CLASSIFICATION SYSTEM"
    ws_main['A1'].font = Font(bold=True, size=16)
    ws_main.merge_cells('A1:M1')
    
    # Basic Info Section
    ws_main['A3'] = "BASIC INFORMATION"
    ws_main['A3'].font = subheader_font
    ws_main['A3'].fill = subheader_fill
    ws_main.merge_cells('A3:D3')
    
    ws_main['A4'] = "Opponent Name:"
    ws_main['B4'] = ""  # Input cell
    ws_main['C4'] = "Match Date:"
    ws_main['D4'] = ""  # Input cell
    
    ws_main['A5'] = "Competition:"
    ws_main['B5'] = ""  # Input cell
    ws_main['C5'] = "Venue:"
    ws_main['D5'] = ""  # Input cell (Home/Away/Neutral)
    
    # Scoring Factors Section
    ws_main['A7'] = "SCORING FACTORS"
    ws_main['A7'].font = subheader_font
    ws_main['A7'].fill = subheader_fill
    ws_main.merge_cells('A7:M7')
    
    # Headers for scoring table
    headers = ['Factor', 'Raw Data', 'Score (0-1)', 'Weight (%)', 'Weighted Points', 'Max Points']
    for i, header in enumerate(headers, 1):
        cell = ws_main.cell(row=8, column=i)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Scoring factors data
    factors = [
        ['1. League Position', '', '=IF(B9="","",1-(B9-1)/19)', '25%', '=C9*0.25', '25'],
        ['2. Recent Form (Points)', '', '=IF(B10="","",B10/15)', '20%', '=C10*0.20', '20'],
        ['3. Home/Away Advantage', '', '=IF(B11="Home",1,IF(B11="Away",0.8,IF(B11="Neutral",0.9,"")))', '15%', '=C11*0.15', '15'],
        ['4. Head-to-Head Record', '', '', '15%', '=C12*0.15', '15'],
        ['5. Key Players Available', '', '=IF(B13="","",B13/11)', '10%', '=C13*0.10', '10'],
        ['6. Tactical Complexity', '', '', '10%', '=C14*0.10', '10'],
        ['7. Physical Intensity', '', '', '5%', '=C15*0.05', '5']
    ]
    
    for i, factor in enumerate(factors, 9):
        for j, value in enumerate(factor, 1):
            cell = ws_main.cell(row=i, column=j)
            cell.value = value
            cell.border = border
            if j == 1:  # Factor name
                cell.font = Font(bold=True)
            elif j in [3, 5]:  # Formula columns
                cell.alignment = center_align
    
    # Total calculation
    ws_main['A16'] = "TOTAL SCORE"
    ws_main['A16'].font = Font(bold=True, size=12)
    ws_main['E16'] = "=SUM(E9:E15)"
    ws_main['E16'].font = Font(bold=True, size=12)
    ws_main['F16'] = "100"
    ws_main['F16'].font = Font(bold=True, size=12)
    
    # Final rating
    ws_main['A18'] = "FINAL DIFFICULTY RATING"
    ws_main['A18'].font = subheader_font
    ws_main['A18'].fill = subheader_fill
    ws_main.merge_cells('A18:D18')
    
    ws_main['A19'] = "Points (0-100):"
    ws_main['B19'] = "=E16"
    ws_main['B19'].font = Font(bold=True, size=14)
    
    ws_main['A20'] = "Rating (1-5):"
    ws_main['B20'] = "=IF(E16>=90,5,IF(E16>=80,4+(E16-80)/10,IF(E16>=60,3+(E16-60)/20,IF(E16>=40,2+(E16-40)/20,1+(E16-20)/20))))"
    ws_main['B20'].font = Font(bold=True, size=14, color="FF0000")
    
    ws_main['A21'] = "Classification:"
    ws_main['B21'] = '=IF(E16>=90,"EXTREMELY DIFFICULT",IF(E16>=80,"DIFFICULT",IF(E16>=60,"MODERATE",IF(E16>=40,"EASY","VERY EASY"))))'
    ws_main['B21'].font = Font(bold=True, size=14, color="FF0000")
    
    # Input guides
    ws_main['G9'] = "Enter league position (1-20)"
    ws_main['G10'] = "Enter points from last 5 games (0-15)"
    ws_main['G11'] = "Select: Home, Away, or Neutral"
    ws_main['G12'] = "Enter H2H score manually (0-1)"
    ws_main['G13'] = "Enter number of key players (0-11)"
    ws_main['G14'] = "Enter tactical score manually (0-1)"
    ws_main['G15'] = "Enter physical score manually (0-1)"
    
    # === INSTRUCTIONS SHEET ===
    print("Creating instructions sheet...")
    
    instructions = [
        ["OPPONENT DIFFICULTY CLASSIFICATION SYSTEM", ""],
        ["", ""],
        ["OVERVIEW", ""],
        ["This spreadsheet calculates opponent difficulty on a scale of 1.0-5.0", ""],
        ["based on 7 weighted factors with scientific backing.", ""],
        ["", ""],
        ["HOW TO USE", ""],
        ["1. Fill in basic information (opponent name, date, venue)", ""],
        ["2. Enter raw data for each scoring factor", ""],
        ["3. The spreadsheet automatically calculates scores and final rating", ""],
        ["", ""],
        ["SCORING FACTORS GUIDE", ""],
        ["", ""],
        ["1. LEAGUE POSITION (25% weight)", ""],
        ["   - Enter current league position (1-20)", ""],
        ["   - Formula: 1-(position-1)/19", ""],
        ["   - Example: 1st place = 1.0, 10th place = 0.53", ""],
        ["", ""],
        ["2. RECENT FORM (20% weight)", ""],
        ["   - Count points from last 5 matches", ""],
        ["   - Win = 3 points, Draw = 1 point, Loss = 0 points", ""],
        ["   - Maximum possible = 15 points", ""],
        ["   - Example: 3W-1D-1L = 10 points", ""],
        ["", ""],
        ["3. HOME/AWAY ADVANTAGE (15% weight)", ""],
        ["   - Home = 1.0, Away = 0.8, Neutral = 0.9", ""],
        ["   - Automatically calculated based on venue selection", ""],
        ["", ""],
        ["4. HEAD-TO-HEAD RECORD (15% weight)", ""],
        ["   - Manual entry required (0-1 scale)", ""],
        ["   - Consider last 5-10 encounters", ""],
        ["   - 0.0 = Very poor record, 1.0 = Excellent record", ""],
        ["", ""],
        ["5. KEY PLAYERS AVAILABLE (10% weight)", ""],
        ["   - Enter number of key players available (0-11)", ""],
        ["   - Formula: available_players/11", ""],
        ["   - Example: 9 players available = 0.82", ""],
        ["", ""],
        ["6. TACTICAL COMPLEXITY (10% weight)", ""],
        ["   - Manual entry required (0-1 scale)", ""],
        ["   - 0.2 = Very Low, 0.4 = Low, 0.6 = Medium", ""],
        ["   - 0.8 = High, 1.0 = Very High", ""],
        ["", ""],
        ["7. PHYSICAL INTENSITY (5% weight)", ""],
        ["   - Manual entry required (0-1 scale)", ""],
        ["   - 0.2 = Very Low, 0.4 = Low, 0.6 = Medium", ""],
        ["   - 0.8 = High, 1.0 = Very High", ""],
        ["", ""],
        ["FINAL CLASSIFICATION SCALE", ""],
        ["90+ points = 5.0 (EXTREMELY DIFFICULT)", ""],
        ["80-89 points = 4.0-4.9 (DIFFICULT)", ""],
        ["60-79 points = 3.0-3.9 (MODERATE)", ""],
        ["40-59 points = 2.0-2.9 (EASY)", ""],
        ["<40 points = 1.0-1.9 (VERY EASY)", ""]
    ]
    
    for i, (instruction, _) in enumerate(instructions, 1):
        ws_instructions[f'A{i}'] = instruction
        if instruction in ["OPPONENT DIFFICULTY CLASSIFICATION SYSTEM", "OVERVIEW", "HOW TO USE", "SCORING FACTORS GUIDE", "FINAL CLASSIFICATION SCALE"]:
            ws_instructions[f'A{i}'].font = Font(bold=True, size=12)
    
    # === EXAMPLES SHEET ===
    print("Creating examples sheet...")
    
    examples_data = [
        ["EXAMPLE 1: Real Madrid (Away)", "", "", "", "", ""],
        ["Factor", "Raw Data", "Score", "Weight", "Points", "Calculation"],
        ["League Position", "1", "1.00", "25%", "25.0", "1st place = maximum score"],
        ["Recent Form", "13", "0.87", "20%", "17.4", "4W-1D = 13 points"],
        ["Home/Away", "Away", "0.80", "15%", "12.0", "Away match penalty"],
        ["Head-to-Head", "Poor", "0.40", "15%", "6.0", "Historical disadvantage"],
        ["Key Players", "10", "0.91", "10%", "9.1", "10 of 11 key players"],
        ["Tactical Complexity", "Very High", "1.00", "10%", "10.0", "Complex tactical system"],
        ["Physical Intensity", "High", "0.80", "5%", "4.0", "High pressing team"],
        ["", "", "", "TOTAL:", "83.5", ""],
        ["", "", "", "RATING:", "4.4/5", "DIFFICULT"],
        ["", "", "", "", "", ""],
        ["EXAMPLE 2: Lower League Team (Home)", "", "", "", "", ""],
        ["Factor", "Raw Data", "Score", "Weight", "Points", "Calculation"],
        ["League Position", "18", "0.11", "25%", "2.8", "18th place = low score"],
        ["Recent Form", "4", "0.27", "20%", "5.4", "1W-1D-3L = 4 points"],
        ["Home/Away", "Home", "1.00", "15%", "15.0", "Home advantage"],
        ["Head-to-Head", "Good", "0.70", "15%", "10.5", "Historical advantage"],
        ["Key Players", "8", "0.73", "10%", "7.3", "8 of 11 key players"],
        ["Tactical Complexity", "Low", "0.40", "10%", "4.0", "Simple tactical approach"],
        ["Physical Intensity", "Medium", "0.60", "5%", "3.0", "Moderate intensity"],
        ["", "", "", "TOTAL:", "48.0", ""],
        ["", "", "", "RATING:", "2.4/5", "EASY"]
    ]
    
    for i, row in enumerate(examples_data, 1):
        for j, value in enumerate(row, 1):
            cell = ws_examples.cell(row=i, column=j)
            cell.value = value
            if i in [1, 13] or (i in [2, 14] and j <= 6):
                cell.font = Font(bold=True)
                if i in [1, 13]:
                    cell.fill = subheader_fill
    
    # Adjust column widths - set standard widths
    column_widths = {
        'A': 25, 'B': 20, 'C': 15, 'D': 15, 'E': 15, 'F': 12, 'G': 35
    }
    
    for ws in [ws_main, ws_instructions, ws_examples]:
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
    
    # Save file
    output_path = "Opponent_Difficulty_Classification_Template.xlsx"
    wb.save(output_path)
    
    print(f"✅ Spreadsheet created successfully: {output_path}")
    print(f"📁 Location: {os.path.abspath(output_path)}")
    
    return output_path

if __name__ == "__main__":
    create_opponent_difficulty_spreadsheet()
